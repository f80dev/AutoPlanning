import random
import datetime
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from tools import intersection, union, filter_plage
from _types import Salle,Professeur,Seance,Cours




class Agenda:

    salles:list[Salle]=[]
    professeurs:list[Professeur]=[]
    cours:list[Cours]=[]
    seances:list[Seance]=[]

    # Si vous modifiez ces portées, supprimez le fichier token.json.
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']


    def get_google_creds(self):
        """Authentification Google centralisée, retourne les credentials."""
        scope = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        return Credentials.from_service_account_file("credentials.json")






    def load_data_from_sheet(self,spreadsheet_id, sheet_name):
        """
        Charge les données d'une feuille de calcul Google dans une liste de listes.
        :param spreadsheet_id: L'ID de votre feuille de calcul.
        :param range_name: La plage à lire, par exemple 'Feuille1!A1:B10'.
        :return: Une liste de listes contenant les données, ou None en cas d'erreur.
        """
        try:

            if "/" in spreadsheet_id:
                # La première ligne est automatiquement utilisée comme en-tête.
                df = pd.read_excel(spreadsheet_id, sheet_name=sheet_name)

                # Convertir le DataFrame en une liste de dictionnaires.
                # L'orientation 'records' crée exactement le format souhaité.
                rc = df.to_dict(orient='records')
            else:
                service = build('sheets', 'v4', credentials=self.get_google_creds())

                # Appeler l'API Sheets
                sheet = service.spreadsheets()
                result = sheet.values().get(spreadsheetId=spreadsheet_id,
                                            range=sheet_name).execute()
                values = result.get('values', [])

                cols=values[0]
                rc=[]
                for v in values[1:]:
                    rc.append(dict(zip(cols, v)))

            return rc

        except HttpError as err:
            print(f"Une erreur s'est produite: {err}")
            return None




    def convert_plage(self,d:dict) -> tuple:
        debut=d["dtStart"]
        fin=d["dtEnd"]
        if type(debut)==str:
            debut=debut.replace("  "," ").replace(" ","/2026 ")
            debut=datetime.datetime.strptime(debut,"%d/%m/%Y %H:%M")

        if type(fin)==str:
            fin = fin.replace("  "," ").replace(" ", "/2026 ")
            fin=datetime.datetime.strptime(fin,"%d/%m/%Y %H:%M")

        return (debut,fin)



    def convert_plage_old(self,debut, fin) -> list:
        """
        Génère une liste de créneaux horaires d'une heure entre une date de début et de fin.

        Args:
            debut (datetime): Le moment de début de la plage (inclus).
            fin (datetime): Le moment de fin de la plage (exclus).

        Returns:
            list[datetime]: Une liste d'objets datetime, un pour chaque heure
                            commençant dans l'intervalle. Retourne une liste vide
                            si la date de début est après ou égale à la date de fin.
        """

        if type(debut)==str:
            debut=debut.replace("  "," ").replace(" ","/2026 ")
            debut=datetime.datetime.strptime(debut,"%d/%m/%Y %H:%M")

        if type(fin)==str:
            fin = fin.replace("  "," ").replace(" ", "/2026 ")
            fin=datetime.datetime.strptime(fin,"%d/%m/%Y %H:%M")

        if debut >= fin:
            return []

        heures_disponibles = []
        heure_actuelle = debut

        # On continue tant que l'heure actuelle est strictement inférieure à l'heure de fin
        while heure_actuelle < fin:
            heures_disponibles.append(heure_actuelle)
            # On ajoute une heure pour passer au créneau suivant
            heure_actuelle += datetime.timedelta(hours=1)

        return heures_disponibles

    import random
    from dataclasses import dataclass

    import pandas as pd

    import datetime
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError

    @dataclass
    class Seance:
        dtStart: datetime.datetime
        dtEnd: datetime.datetime
        salle: str
        Prof_ID: str

        def __repr__(self):
            return f"Seance(dtStart='{self.dtStart}', dtEnd='{self.dtEnd}', salle='{self.salle}', Prof_ID='{self.Prof_ID}')"

    @dataclass
    class Cours:
        duree: int
        titre: str
        nb_etudiant: int
        Prof_ID: str
        requirments: str = ""


    @dataclass
    class Professeur:
        Prof_ID: str
        Nom: str
        disponibilites = []

    @dataclass
    class Salle:
        Salle_ID: str
        capacite: int
        tags: str = ""
        disponibilites = []

    class Agenda:

        salles = []
        professeurs = []
        cours = []
        seances = []

        # Si vous modifiez ces portées, supprimez le fichier token.json.
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

        def get_google_creds(self):
            """Authentification Google centralisée, retourne les credentials."""
            scope = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            return Credentials.from_service_account_file("credentials.json")

        def load_data_from_sheet(self, spreadsheet_id, sheet_name):
            """
            Charge les données d'une feuille de calcul Google dans une liste de listes.
            :param spreadsheet_id: L'ID de votre feuille de calcul.
            :param range_name: La plage à lire, par exemple 'Feuille1!A1:B10'.
            :return: Une liste de listes contenant les données, ou None en cas d'erreur.
            """
            try:

                if "/" in spreadsheet_id:
                    # La première ligne est automatiquement utilisée comme en-tête.
                    df = pd.read_excel(spreadsheet_id, sheet_name=sheet_name)

                    # Convertir le DataFrame en une liste de dictionnaires.
                    # L'orientation 'records' crée exactement le format souhaité.
                    rc = df.to_dict(orient='records')
                else:
                    service = build('sheets', 'v4', credentials=self.get_google_creds())

                    # Appeler l'API Sheets
                    sheet = service.spreadsheets()
                    result = sheet.values().get(spreadsheetId=spreadsheet_id,
                                                range=sheet_name).execute()
                    values = result.get('values', [])

                    cols = values[0]
                    rc = []
                    for v in values[1:]:
                        rc.append(dict(zip(cols, v)))

                return rc

            except HttpError as err:
                print(f"Une erreur s'est produite: {err}")
                return None

        def convert_plage(self, debut, fin) -> tuple:
            if type(debut) == str:
                debut = debut.replace("  ", " ").replace(" ", "/2026 ")
                debut = datetime.datetime.strptime(debut, "%d/%m/%Y %H:%M")

            if type(fin) == str:
                fin = fin.replace("  ", " ").replace(" ", "/2026 ")
                fin = datetime.datetime.strptime(fin, "%d/%m/%Y %H:%M")

            return (debut, fin)

        def convert_plage_old(self, debut, fin) -> list:
            """
            Génère une liste de créneaux horaires d'une heure entre une date de début et de fin.

            Args:
                debut (datetime): Le moment de début de la plage (inclus).
                fin (datetime): Le moment de fin de la plage (exclus).

            Returns:
                list[datetime]: Une liste d'objets datetime, un pour chaque heure
                                commençant dans l'intervalle. Retourne une liste vide
                                si la date de début est après ou égale à la date de fin.
            """

            if type(debut) == str:
                debut = debut.replace("  ", " ").replace(" ", "/2026 ")
                debut = datetime.datetime.strptime(debut, "%d/%m/%Y %H:%M")

            if type(fin) == str:
                fin = fin.replace("  ", " ").replace(" ", "/2026 ")
                fin = datetime.datetime.strptime(fin, "%d/%m/%Y %H:%M")

            if debut >= fin:
                return []

            heures_disponibles = []
            heure_actuelle = debut

            # On continue tant que l'heure actuelle est strictement inférieure à l'heure de fin
            while heure_actuelle < fin:
                heures_disponibles.append(heure_actuelle)
                # On ajoute une heure pour passer au créneau suivant
                heure_actuelle += datetime.timedelta(hours=1)

            return heures_disponibles

        # "1e7wdfc2brpNwbefhA9dJXx81zs93bGpRM_Th_eiHwBI"
        def init_listes(self, classeur_id="./planning.xlsx"):
            self.professeurs = []
            for p in self.load_data_from_sheet(classeur_id, "Professeurs"):
                self.professeurs.append(Professeur(**p))

            self.cours = [Cours(**p) for p in self.load_data_from_sheet(classeur_id, "Cours")]
            self.salles = [Salle(**p) for p in self.load_data_from_sheet(classeur_id, "Salles")]

            dispo_prof = self.load_data_from_sheet(classeur_id, "DispoProf")
            for p in self.professeurs:
                p.disponibilites.clear()
                for d in dispo_prof:
                    if p.Prof_ID == d["Prof_ID"]:
                        p.disponibilites.append(self.convert_plage(d["dtStart"], d["dtEnd"]))

            dispo_salle = self.load_data_from_sheet(classeur_id, "DispoSalle")
            for s in self.salles:
                s.disponibilites.clear()
                for d in dispo_salle:
                    if s.Salle_ID == d["Salle_ID"]:
                        s.disponibilites.append(self.convert_plage(d["dtStart"], d["dtEnd"]))

        def get_salle(self, min_capacite=0) -> Salle | None:
            if min_capacite == 0:
                return random.choice(self.salles)
            else:
                for _ in range(1000):
                    s = random.choice(self.salles)
                    if s.capacite >= min_capacite: return s
                return None

        def get_cours(self) -> Cours | None:
            if len(self.cours) == 0: return None
            return random.choice(self.cours)

        def check_plage(self, disponibilite: list, plage: tuple) -> tuple | None:
            for d in disponibilite:
                if d[0] <= plage[0] and d[1] >= plage[1]: return d
            return None

        def get_prof(self, prof_id=None):
            if prof_id is None: return random.choice(self.professeurs)
            for p in self.professeurs:
                if p.Prof_ID == prof_id: return p
            return None

        def update_plage(self, dispo: tuple, plage: tuple) -> tuple | None:
            """
            reserve une plage dans une disponibilité (soit au début soit à la fin)
            :param dispo:
            :param plage:
            :return: la disponibilité
            """
            if dispo[0] > plage[0] or dispo[1] < plage[1]: return None
            if random.choice([0, 1]) == 0:
                return (plage[1], dispo[1])
            else:
                return (dispo[0], plage[0])

        def convert_duration_to_plage(self, dispo: list, duree: int = 1):
            plage = self.find_plage_for_duration(dispo, duree)
            if plage is None: return None
            return (plage[0], plage[0] + datetime.timedelta(hours=duree))

        def update_dispos(self, dispo: list, plage: tuple):
            p = self.check_plage(dispo, plage)
            dispo.append(self.update_plage(p, plage))
            dispo.remove(p)
            return dispo

        def reserve(self, dispo: list, plage: tuple):
            for d in dispo:
                if d[0] <= plage[0] and d[1] >= plage[1]:
                    if d[0] != plage[0]: dispo.append((d[0], plage[0]))
                    if plage[1] != d[1]: dispo.append((plage[1], d[1]))
                    dispo.remove(d)
                    return dispo

            raise RuntimeError("La réservation ne fonctionne pas, c'est anormal")

        def find_plage_for_duration(self, disponibilite: list, duree: int) -> tuple | None:
            """
            cherche une disponiblité et met a jour les dispos suite à réservation
            :param disponibilite:
            :param duree:
            :return:
            """
            for _ in range(10000):
                # on tire au sort dans les disponibilite
                d = random.choice(disponibilite)
                if (d[1] - d[0]).total_seconds() >= duree * 3600:
                    return d

            return None




    #"1e7wdfc2brpNwbefhA9dJXx81zs93bGpRM_Th_eiHwBI"
    def init_listes(self,classeur_id ="./planning.xlsx" ):
        self.professeurs = [Professeur(**p) for p in self.load_data_from_sheet(classeur_id, "Professeurs")]
        self.cours = [Cours(**p) for p in self.load_data_from_sheet(classeur_id, "Cours")]
        self.salles = [Salle(**p) for p in self.load_data_from_sheet(classeur_id, "Salles")]

        dispo_prof=self.load_data_from_sheet(classeur_id, "DispoProf")
        for p in self.professeurs:
            p.disponibilites.clear()
            for d in dispo_prof:
                if p.Prof_ID==d["Prof_ID"]:
                    p.disponibilites.append(self.convert_plage(d))

            p.disponibilites=union(p.disponibilites)

        dispo_salle = self.load_data_from_sheet(classeur_id, "DispoSalle")
        for s in self.salles:
            s.tags="" if type(s.tags)==float else s.tags
            s.disponibilites.clear()
            for d in dispo_salle:
                if s.Salle_ID==d["Salle_ID"]:
                    s.disponibilites.append(self.convert_plage(d))

            s.disponibilites = union(s.disponibilites)

        for c in self.cours:
            c.requirments="" if type(c.requirments)==float else c.requirments


        return {"professeurs":self.professeurs,"cours":self.cours,"salles":self.salles}


    def get_salle(self,min_capacite=0) -> Salle | None:
        if min_capacite==0:
            return random.choice(self.salles)
        else:
            for _ in range(1000):
                s=random.choice(self.salles)
                if s.capacite>=min_capacite: return s
            return None




    def get_cours(self) -> Cours | None:
        if len(self.cours)==0: return None
        return random.choice(self.cours)


    def check_plage(self,disponibilite:list,plage:tuple) -> tuple | None:
        for d in disponibilite:
            if d[0]<=plage[0] and d[1]>=plage[1]:return d
        return None


    def get_prof(self,prof_id=None):
        if prof_id is None: return random.choice(self.professeurs)
        for p in self.professeurs:
            if p.Prof_ID==prof_id: return p
        return None


    def update_plage(self,dispo:tuple,plage:tuple) -> tuple | None:
        """
        reserve une plage dans une disponibilité (soit au début soit à la fin)
        :param dispo:
        :param plage:
        :return: la disponibilité
        """
        if dispo[0]>plage[0] or dispo[1]<plage[1]: return None
        if random.choice([0,1])==0:
            return (plage[1],dispo[1])
        else:
            return (dispo[0],plage[0])


    def convert_duration_to_plage(self,dispo:list,duree:int=1):
        plage=self.find_plage_for_duration(dispo,duree)
        if plage is None: return None
        return (plage[0], plage[0]+datetime.timedelta(hours=duree))


    def update_dispos(self,dispo:list,plage:tuple):
        p=self.check_plage(dispo,plage)
        dispo.append(self.update_plage(p,plage))
        dispo.remove(p)
        return dispo


    def reserve(self,dispo:list,plage:tuple):
        for d in dispo:
            if d[0]<=plage[0] and d[1]>=plage[1]:
                if d[0]!=plage[0]:dispo.append((d[0],plage[0]))
                if plage[1]!=d[1]:dispo.append((plage[1],d[1]))
                dispo.remove(d)
                return dispo

        raise RuntimeError("La réservation ne fonctionne pas, c'est anormal")




    def find_plage_for_duration(self,disponibilite:list, duree:int) -> tuple | None:
        """
        cherche une disponiblité et met a jour les dispos suite à réservation
        :param disponibilite:
        :param duree:
        :return:
        """
        for _ in range(10000):
            #on tire au sort dans les disponibilite
            d=random.choice(disponibilite)
            if (d[1]-d[0]).total_seconds()>=duree*3600:
                return d

        return None

    def run(self,filename="./planning.xlsx") -> list[Seance] | None:
        for occ in range(1000):
            rc = False
            self.init_listes(filename)

            planning = []
            total_cours = len(self.cours)

            for i in range(1000):
                c = self.get_cours()
                if c is None:
                    rc = True
                    break

                p = self.get_prof(c.Prof_ID)
                s = self.get_salle(min_capacite=c.nb_etudiant)

                if c.nb_etudiant <= s.capacite and (c.requirments=="" or set(c.requirments.split(",")).issubset(s.tags.split(","))):
                    plage_seance = self.convert_duration_to_plage(s.disponibilites, c.duree)
                    if plage_seance:
                        if self.check_plage(p.disponibilites, plage_seance):
                            planning.append(Seance(plage_seance[0], plage_seance[1], s.Salle_ID, p.Prof_ID, c.titre, p.Nom))
                            s.disponibilites = a.reserve(s.disponibilites, plage_seance)
                            p.disponibilites = a.reserve(p.disponibilites, plage_seance)
                            self.cours.remove(c)

            if rc:
                return(planning)
            else:
                print(f"Echec de planification {occ} tentatives. {total_cours - len(self.cours)}/{total_cours} planifiés")

        return None





def export_planning_to_excel(planning: list[Seance], filename: str = "planning_output.xlsx", min_hour: int = 0, max_hour: int = 23):
    """
    Exporte le planning dans un fichier Excel avec un onglet visuel et un onglet de données.
    """
    if not planning:
        print("Le planning est vide, aucun fichier n'a été généré.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Onglet Planning Visuel ---

    # 1. Déterminer la plage de dates et les salles
    min_date = min(s.dtStart for s in planning).replace(minute=0, second=0, microsecond=0)
    max_date = max(s.dtEnd for s in planning)
    salles = sorted(list(set(s.salle for s in planning)))
    
    # Créer un mapping des salles vers les lignes
    salle_to_row = {salle: i + 4 for i, salle in enumerate(salles)}
    for salle, row_idx in salle_to_row.items():
        cell = ws.cell(row=row_idx, column=1, value=salle)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 2. Créer l'en-tête sur 3 niveaux
    ws.cell(row=1, column=1, value="Mois")
    ws.cell(row=2, column=1, value="Jour")
    ws.cell(row=3, column=1, value="Heure")
    
    # Remplir les colonnes de temps, en filtrant par min_hour et max_hour
    time_cols = []
    current_time = min_date
    while current_time < max_date:
        if min_hour <= current_time.hour <= max_hour:
            time_cols.append(current_time)
        current_time += datetime.timedelta(hours=1)

    col_idx = 2
    datetime_to_col = {}
    for dt in time_cols:
        ws.cell(row=1, column=col_idx, value=dt.strftime("%B %Y")).alignment = center_align
        ws.cell(row=2, column=col_idx, value=dt.strftime("%a %d")).alignment = center_align
        ws.cell(row=3, column=col_idx, value=dt.strftime("%H")).alignment = center_align
        datetime_to_col[dt] = col_idx
        col_idx += 1

    # 3. Fusionner les cellules de l'en-tête
    def merge_header_cells(row_idx):
        if ws.max_column < 2: return
        start_col = 2
        for col in range(3, ws.max_column + 2):
            current_val = ws.cell(row=row_idx, column=col).value
            prev_val = ws.cell(row=row_idx, column=col - 1).value
            if current_val != prev_val:
                ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=col - 1)
                start_col = col
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=ws.max_column)

    merge_header_cells(1) # Fusionner les mois
    merge_header_cells(2) # Fusionner les jours

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

    # 4. Remplir le planning avec les séances
    for seance in planning:
        if seance.salle not in salle_to_row: continue
        
        row_idx = salle_to_row[seance.salle]
        start_dt = seance.dtStart.replace(minute=0, second=0, microsecond=0)
        end_dt = seance.dtEnd.replace(minute=0, second=0, microsecond=0)
        
        start_col = datetime_to_col.get(start_dt)
        end_col_dt = end_dt - datetime.timedelta(hours=1)
        if end_col_dt < start_dt: end_col_dt = start_dt
        end_col = datetime_to_col.get(end_col_dt)

        if start_col and end_col:
            ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
            cell = ws.cell(row=row_idx, column=start_col)
            cell.value = f"{seance.titre} - {seance.Nom_Prof}"
            cell.alignment = center_align
            cell.border = thin_border

    # 5. Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 12
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 8

    # --- Onglet Séances ---
    ws_seances = wb.create_sheet("Seances")
    
    # En-têtes
    seance_headers = ["Date Début", "Date Fin", "Salle", "Titre", "Professeur", "ID Prof"]
    ws_seances.append(seance_headers)
    for cell in ws_seances[1]:
        cell.font = Font(bold=True)

    # Remplissage des données
    for seance in planning:
        row_data = [
            seance.dtStart,
            seance.dtEnd,
            seance.salle,
            seance.titre,
            seance.Nom_Prof,
            seance.Prof_ID
        ]
        ws_seances.append(row_data)
        
    # Ajuster la largeur des colonnes pour l'onglet séances
    for i, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F']):
        ws_seances.column_dimensions[col_letter].width = 20


    # --- Sauvegarde ---
    try:
        wb.save(filename)
        print(f"Le planning a été exporté avec succès dans '{filename}'")
    except Exception as e:
        print(f"Erreur lors de la sauvegarde du fichier Excel : {e}")


if __name__ == '__main__':
    a=Agenda()
    planning=a.run()
    if planning:
        export_planning_to_excel(
            filter_plage(planning,datetime.datetime.now(),datetime.datetime.now()+datetime.timedelta(days=30)),
            min_hour=8,
            max_hour=20
        )
