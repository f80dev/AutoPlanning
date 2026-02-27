import datetime
import json
from dataclasses import asdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from _types import Seance, Plage


def custom_serializer(obj):
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} not serializable")




def export_planning_to_json(planning: list[Seance], filename: str = "planning_output.json"):
    if not planning:
        print("Le planning est vide, aucun fichier n'a été généré.")
        return

    data = [asdict(s) for s in planning]

    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False, default=custom_serializer)
    print(f"Le planning a été exporté avec succès dans '{filename}'")




def export_plages_to_json(planning: list[Plage], filename: str = "planning_output.json",titre="Plage"):
    if not planning:
        print("Le planning est vide, aucun fichier n'a été généré.")
        return

    data=[]
    for i,p in enumerate(planning):
        d={
            "Prof_ID":"",
            "titre":titre+str(i),
            "props":{},
            "group":"",
            "plage":p
        }
        data.append(d)

    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False, default=custom_serializer)
    print(f"Le planning a été exporté avec succès dans '{filename}'")



def export_planning_to_excel(planning: list[Seance], filename: str = "planning_output.xlsx", min_hour: int = 0,
                             max_hour: int = 23):
    """
    Exporte le planning dans un fichier Excel avec un onglet visuel et un onglet de données.
    """
    if not planning:
        print("Le planning est vide, aucun fichier n'a été généré.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = (PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"))
    cell_fill = (PatternFill(start_color="40FFFF", end_color="40FFFF", fill_type="solid"))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # --- Onglet Planning Visuel ---

    # 1. Déterminer la plage de dates et les salles
    min_date = min(s.plage.dtStart for s in planning).replace(minute=0, second=0, microsecond=0)
    max_date = max(s.plage.dtEnd for s in planning)
    salles = sorted(list(set(s.salle for s in planning)))

    # Créer un mapping des salles vers les lignes
    salle_to_row = {salle: i + 4 for i, salle in enumerate(salles)}
    for salle, row_idx in salle_to_row.items():
        cell = ws.cell(row=row_idx, column=1, value=salle)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 2. Créer l'en-tête sur 3 niveaux
    ws.cell(row=1, column=1, value="Mois")
    ws.cell(row=2, column=1, value="Jour")
    ws.cell(row=3, column=1, value="Heure")

    # Remplir les colonnes de temps, en filtrant par min_hour et max_hour
    time_cols = []
    current_time = min_date
    while current_time < max_date:
        if min_hour <= current_time.hour <= max_hour:
            time_cols.append(current_time)
        current_time += datetime.timedelta(hours=1)

    col_idx = 2
    datetime_to_col = {}
    for dt in time_cols:
        ws.cell(row=1, column=col_idx, value=dt.strftime("%B %Y")).alignment = center_align
        ws.cell(row=2, column=col_idx, value=dt.strftime("%a %d")).alignment = center_align
        ws.cell(row=3, column=col_idx, value=dt.strftime("%H")).alignment = center_align
        datetime_to_col[dt] = col_idx
        col_idx += 1

    # 3. Fusionner les cellules de l'en-tête
    def merge_header_cells(row_idx):
        if ws.max_column < 2: return
        start_col = 2
        for col in range(3, ws.max_column + 2):
            current_val = ws.cell(row=row_idx, column=col).value
            prev_val = ws.cell(row=row_idx, column=col - 1).value
            if current_val != prev_val:
                ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=col - 1)
                start_col = col
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=ws.max_column)

    merge_header_cells(1)  # Fusionner les mois
    merge_header_cells(2)  # Fusionner les jours

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

    # 4. Remplir le planning avec les séances
    for seance in planning:
        if seance.salle not in salle_to_row: continue

        row_idx = salle_to_row[seance.salle]
        start_dt = seance.plage.dtStart.replace(minute=0, second=0, microsecond=0)
        end_dt = seance.plage.dtEnd.replace(minute=0, second=0, microsecond=0)

        start_col = datetime_to_col.get(start_dt)
        end_col_dt = end_dt - datetime.timedelta(hours=1)
        if end_col_dt < start_dt: end_col_dt = start_dt
        end_col = datetime_to_col.get(end_col_dt)

        if start_col and end_col:
            ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
            cell = ws.cell(row=row_idx, column=start_col)
            cell.value = f"{seance.titre} - Groupe {seance.group} - {seance.Nom_Prof}"
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill=cell_fill

    # 5. Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 12
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 8

    # --- Onglet Séances ---
    ws_seances = wb.create_sheet("Seances")

    # En-têtes
    seance_headers = ["Date Début", "Date Fin", "Salle", "Titre", "Professeur", "ID Prof"]
    ws_seances.append(seance_headers)
    for cell in ws_seances[1]:
        cell.font = Font(bold=True)

    # Remplissage des données
    for seance in planning:
        row_data = [
            seance.plage.dtStart,
            seance.plage.dtEnd,
            seance.salle,
            seance.titre,
            seance.Nom_Prof,
            seance.Prof_ID
        ]
        ws_seances.append(row_data)

    # Ajuster la largeur des colonnes pour l'onglet séances
    for i, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F']):
        ws_seances.column_dimensions[col_letter].width = 20

    # --- Sauvegarde ---
    try:
        wb.save(filename)
        print(f"Le planning a été exporté avec succès dans '{filename}'")
    except Exception as e:
        print(f"Erreur lors de la sauvegarde du fichier Excel : {e}")
